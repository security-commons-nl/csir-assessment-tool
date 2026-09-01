"""csir.json tegen de werkboeken.

De pagina is niets waard als de tekst erin niet die van de richtlijn is. Deze tests lezen de twee
werkboeken zelfstandig met openpyxl en leggen elke tekst, elk nummer en elke niveaumarkering naast
csir.json. Faalt hier iets, repareer dan de bron of draai haal_bron.py opnieuw; nooit de test.
"""
from __future__ import annotations

import hashlib
import json
import pathlib
import re
import subprocess
import sys

import openpyxl
import pytest

ROOT = pathlib.Path(__file__).resolve().parent.parent.parent
WERKBOEK = ROOT / "werkboek" / "csir-control-register.xlsx"
CLASSIFICATIE = ROOT / "werkboek" / "objectclassificatie.xlsx"

VERWACHT_PER_PARAGRAAF = {
    "§2.1.1": 32, "§2.1.2": 33, "§2.2": 31, "§2.3": 15, "§2.4.1": 16, "§2.4.2": 7,
    "§2.5.1": 10, "§2.5.2": 10, "§2.5.3": 8, "§2.6": 33, "§2.7.1": 27, "§2.7.2": 6,
    "§2.8": 15, "§2.9": 18, "§2.10": 7,
}
VERWACHT_PER_NIVEAU = {1: 193, 2: 198, 3: 230, 4: 234}
CRITERIA = [("veiligheid", "E", "F"), ("maatschappij", "G", "H"), ("financieel", "I", "J"),
            ("cascade", "K", "L"), ("ecologie", "M", "N"), ("imago", "O", "P")]


@pytest.fixture(scope="module")
def wb():
    return openpyxl.load_workbook(WERKBOEK, data_only=True)


@pytest.fixture(scope="module")
def wb_formules():
    return openpyxl.load_workbook(WERKBOEK, data_only=False)


@pytest.fixture(scope="module")
def wb_klas():
    return openpyxl.load_workbook(CLASSIFICATIE, data_only=True)


def celtekst(ws, rij, kolom) -> str:
    waarde = ws.cell(rij, kolom).value
    if waarde is None:
        return ""
    if isinstance(waarde, float) and waarde.is_integer():
        return str(int(waarde))
    return str(waarde).strip()


def test_aantallen(bron):
    assert sum(1 for c in bron["controls"] if c["blad"] == "VSP") == 89
    assert sum(1 for c in bron["controls"] if c["blad"] == "VSE") == 38
    assert len(bron["maatregelen"]) == 268
    assert len(bron["bijlagen"]) == 27
    assert len(bron["paragrafen"]) == 15
    assert len(bron["classificatie"]["criteria"]) == 6
    assert len(bron["classificatie"]["ernst"]) == 5
    for criterium in bron["classificatie"]["criteria"]:
        assert sorted(criterium["drempels"]) == ["1", "2", "3", "4", "5"]


def test_maatregelen_per_niveau(bron):
    gevonden = {n: sum(1 for m in bron["maatregelen"] if n in m["niveaus"]) for n in (1, 2, 3, 4)}
    assert gevonden == VERWACHT_PER_NIVEAU


def test_maatregelen_per_paragraaf(bron):
    gevonden = {p["id"]: p["maatregelen"] for p in bron["paragrafen"]}
    assert gevonden == VERWACHT_PER_PARAGRAAF
    for paragraaf in bron["paragrafen"]:
        echt = sum(1 for m in bron["maatregelen"] if m["paragraaf"] == paragraaf["id"])
        assert echt == paragraaf["maatregelen"]


def test_control_eisen_woordelijk(bron, wb, wb_formules):
    """Elke eis, korte omschrijving en BIO-bron staat letterlijk zoals in het werkboek.

    Twee cellen op VSE zijn een sjabloonformule: de richtlijntekst met de objectnaam en het niveau
    erin. Daarvan controleren we dat elk letterlijk tekstdeel uit de formule in de bron staat en dat
    de twee plaatshouders open zijn gebleven.
    """
    per_id = {c["id"]: c for c in bron["controls"]}
    for blad, kort, laatste in (("VSP Proceseisen", "VSP", 92), ("VSE Systeemeisen", "VSE", 41)):
        ws = wb[blad]
        ws_f = wb_formules[blad]
        for rij in range(4, laatste + 1):
            nummer = int(float(celtekst(ws, rij, 1)))
            control = per_id[f"{kort}-{nummer}"]
            assert control["bio_bron"] == celtekst(ws, rij, 2)
            for kolom, veld in ((3, "eis"), (11, "kort")):
                ruw = ws_f.cell(rij, kolom).value
                if isinstance(ruw, str) and ruw.startswith("="):
                    assert control["sjabloon"] is True
                    for stuk in re.findall(r'"([^"]*)"', ruw):
                        if stuk.strip():
                            assert stuk.strip() in control[veld], f"{blad} rij {rij}: {stuk[:40]}"
                    assert "{object}" in control[veld] and "{niveau}" in control[veld]
                else:
                    assert control[veld] == celtekst(ws, rij, kolom), f"{blad} rij {rij} {veld}"


def test_maatregelteksten_woordelijk(bron, wb):
    ws = wb["Maatregelen"]
    per_code = {m["code"]: m for m in bron["maatregelen"]}
    for rij in range(4, 272):
        code = celtekst(ws, rij, 4)
        maatregel = per_code[code]
        assert maatregel["paragraaf"] == celtekst(ws, rij, 1)
        assert maatregel["thema"] == celtekst(ws, rij, 2)
        assert maatregel["groep"] == celtekst(ws, rij, 3)
        assert maatregel["tekst"] == celtekst(ws, rij, 5)
        verwacht = [n for n, kolom in enumerate((6, 7, 8, 9), start=1)
                    if celtekst(ws, rij, kolom) == "X"]
        assert maatregel["niveaus"] == verwacht, f"rij {rij}: niveaus van {code}"


def test_bijlagen_woordelijk(bron, wb):
    ws = wb["Bijlagen"]
    per_id = {b["id"]: b for b in bron["bijlagen"]}
    for rij in range(4, 31):
        bijlage = per_id[celtekst(ws, rij, 1)]
        assert bijlage["titel"] == celtekst(ws, rij, 2)
        assert bijlage["type"] == celtekst(ws, rij, 3)
        ruw = celtekst(ws, rij, 4)
        verwacht = int(float(ruw)) if ruw.replace(".", "", 1).isdigit() else None
        assert bijlage["aangeroepen_door"] == verwacht


def test_classificatie_woordelijk(bron, wb_klas):
    ws = wb_klas.active
    for (ident, kolom_uitleg, kolom_drempel), criterium in zip(
            CRITERIA, bron["classificatie"]["criteria"]):
        u = openpyxl.utils.column_index_from_string(kolom_uitleg)
        d = openpyxl.utils.column_index_from_string(kolom_drempel)
        assert criterium["id"] == ident
        assert criterium["titel"] == celtekst(ws, 19, u)
        assert criterium["uitleg"] == celtekst(ws, 20, u)
        for score in range(1, 6):
            assert criterium["drempels"][str(score)] == celtekst(ws, 20 + score, d)

    for rij, ernst in zip(range(21, 26), bron["classificatie"]["ernst"]):
        assert ernst["score"] == int(float(celtekst(ws, rij, 2)))
        assert ernst["label"] == celtekst(ws, rij, 3)
        assert ernst["functiebox"] == celtekst(ws, rij, 18)
        assert ernst["niveau"] == int(float(celtekst(ws, rij, 19)))


def test_sleutels_uniek(bron):
    ids = [c["id"] for c in bron["controls"]]
    codes = [m["code"] for m in bron["maatregelen"]]
    bijlagen = [b["id"] for b in bron["bijlagen"]]
    assert len(set(ids)) == len(ids)
    assert len(set(codes)) == len(codes)
    assert len(set(bijlagen)) == len(bijlagen)
    assert [c["id"] for c in bron["classificatie"]["criteria"]] == [c[0] for c in CRITERIA]


def test_aangeroepen_paragrafen_bestaan(bron):
    """Elke aanroep wijst naar een bestaande paragraaf of naar de bovenliggende daarvan."""
    geldig = {p["id"] for p in bron["paragrafen"]} | {p["ouder"] for p in bron["paragrafen"]}
    met_aanroep = 0
    for control in bron["controls"]:
        for verwijzing in control["aangeroepen"]:
            assert verwijzing in geldig, f"{control['id']} roept {verwijzing} aan"
        if control["aangeroepen"]:
            met_aanroep += 1
            assert len(control["aangeroepen"]) == 1
    assert met_aanroep == 32
    assert len(bron["controls"]) - met_aanroep == 95


def test_bijlage_verwijzingen_bestaan(bron):
    bekend = {b["id"] for b in bron["bijlagen"]}
    for control in bron["controls"]:
        for verwijzing in control["bijlagen"]:
            assert verwijzing in bekend, f"{control['id']} verwijst naar {verwijzing}"


def test_functiebox_tabel(bron, wb):
    assert bron["functiebox_niveau"] == {"A": 4, "B": 3, "C": 2, "D": 1, "E": 1}
    ws = wb["Instellingen"]
    uit_werkboek = {celtekst(ws, rij, 2): int(float(celtekst(ws, rij, 3))) for rij in range(25, 30)}
    assert bron["functiebox_niveau"] == uit_werkboek
    # De classificatie en de instellingen moeten dezelfde tabel gebruiken, anders geeft de tool bij
    # dezelfde functiebox twee verschillende niveaus.
    uit_classificatie = {e["functiebox"]: e["niveau"] for e in bron["classificatie"]["ernst"]}
    assert uit_classificatie == bron["functiebox_niveau"]


def test_keuzelijsten_komen_uit_het_werkboek(bron):
    assert bron["keuzes"]["van_toepassing"] == ["Ja", "Nee", "N.v.t. (buiten scope)"]
    assert bron["keuzes"]["status"] == ["Nog te doen", "In uitvoering", "Geïmplementeerd",
                                        "Explain (afwijking)", "N.v.t."]
    assert bron["keuzes"]["functiebox"] == ["A", "B", "C", "D", "E"]


def test_vingerafdrukken_kloppen(bron):
    """De sha256 van de werkboeken staat in de bron, zodat een stille wijziging opvalt."""
    assert bron["bron"]["werkboek_sha256"] == hashlib.sha256(WERKBOEK.read_bytes()).hexdigest()
    assert bron["bron"]["classificatie_sha256"] == \
        hashlib.sha256(CLASSIFICATIE.read_bytes()).hexdigest()


def test_auteursrecht_staat_erin(bron):
    tekst = bron["bron"]["auteursrecht"]
    assert "Rijkswaterstaat" in tekst
    assert "Het Waterschapshuis" in tekst
    assert "EUPL" in tekst


def test_barrieres_bestaan():
    """Elke barriere in de koppeling kent de kennisbank, met of zonder handleiding."""
    koppeling = json.loads(
        (ROOT / "register" / "paragrafen-barrieres.json").read_text(encoding="utf-8"))
    export = json.loads(
        (ROOT / "register" / "handelingsperspectief.json").read_text(encoding="utf-8"))
    bekend = {h["barriere"] for h in export["handleidingen"]} | set(export["zonder_handleiding"])
    for regel in koppeling["regels"]:
        for barriere in regel["barrieres"]:
            assert barriere in bekend, f"{regel['paragraaf']} noemt onbekende barriere {barriere}"
        assert regel["reden"].strip(), f"{regel['paragraaf']} heeft geen reden"


def test_elke_paragraaf_staat_in_de_koppeling(bron):
    """Stilte is nooit een vergissing: elke paragraaf staat erin, ook met een lege lijst."""
    koppeling = json.loads(
        (ROOT / "register" / "paragrafen-barrieres.json").read_text(encoding="utf-8"))
    genoemd = [r["paragraaf"] for r in koppeling["regels"]]
    assert genoemd == [p["id"] for p in bron["paragrafen"]]


def test_haal_bron_check_slaagt():
    uitkomst = subprocess.run([sys.executable, "register/haal_bron.py", "--check"],
                              cwd=ROOT, capture_output=True, text=True, encoding="utf-8")
    assert uitkomst.returncode == 0, uitkomst.stdout + uitkomst.stderr
