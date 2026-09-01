#!/usr/bin/env python3
"""Haalt de bron uit de twee werkboeken en schrijft csir.json.

De richtlijn zelf is de waarheid; de werkboeken zijn de plek waar die tekst is ingetypt en
gecontroleerd. Dit script haalt hem daaruit naar een vorm die een pagina kan lezen. Vanaf dat moment
is `csir.json` de bron voor het instrument en zijn de werkboeken een export: wie een eistekst
corrigeert, doet dat in het werkboek en draait dit script opnieuw, of corrigeert csir.json en houdt
het werkboek bij. Wat niet mag is de twee stil uit elkaar laten lopen; `--check` en test_bron.py
blokkeren daarop.

Aanroep:
    python register/haal_bron.py            # schrijft csir.json
    python register/haal_bron.py --check    # meldt alleen of csir.json nog klopt (exit 1 als niet)

Vereist openpyxl. Alleen nodig voor dit script en voor test_bron.py; de pagina en het bouwscript
hebben het niet nodig.
"""
from __future__ import annotations

import hashlib
import json
import pathlib
import re
import sys

import openpyxl

HIER = pathlib.Path(__file__).resolve().parent
REPO = HIER.parent
WERKBOEK = REPO / "werkboek" / "csir-control-register.xlsx"
CLASSIFICATIE = REPO / "werkboek" / "objectclassificatie.xlsx"
DOEL = REPO / "csir.json"

VERSIE = "2026-09"

# De zes gevolgcriteria van het classificatieformulier. Per criterium: de kolom met de uitleg (rij 20)
# en de kolom met de vijf drempels (rij 21 tot en met 25). De titels staan in rij 19 boven de uitleg.
CRITERIA = [
    ("veiligheid", "E", "F"),
    ("maatschappij", "G", "H"),
    ("financieel", "I", "J"),
    ("cascade", "K", "L"),
    ("ecologie", "M", "N"),
    ("imago", "O", "P"),
]

# Objecttypen die het formulier niet kent omdat het uit de waterschapswereld komt. De gemeentelijke
# objecten met industriele automatisering horen er net zo goed in; ze gaan voor de vangnetoptie.
EXTRA_HOOFDTAKEN = ["Tunnel", "Gemaal", "Brug", "Sluis", "Verkeersregelinstallatie", "Parkeergarage"]
VANGNET = "Overige (Beschrijf bij opmerking)"

PARAGRAAF = re.compile(r"^§2\.\d+(\.\d+)?$")

AUTEURSRECHT = (
    "De eisteksten, maatregelteksten, niveaumarkeringen en drempels zijn woordelijk overgenomen uit "
    "de Cybersecurity Implementatierichtlijn Objecten (CSIR); het auteursrecht daarop ligt bij "
    "Rijkswaterstaat en Het Waterschapshuis. EUPL-1.2 dekt alleen de opzet van dit register: de "
    "indeling, de rekenregels, de pagina en de documentatie."
)

TOELICHTING_REKENREGEL = (
    "Het formulier neemt het afgeronde gemiddelde van de zes scores (ROUND(som/6;0), half naar "
    "boven). De tool toont daarnaast de strengste lezing (de hoogste score). Welke lezing de CSIR "
    "bedoelt is een open vraag aan de opsteller; tot die is beantwoord volgt de tool het formulier."
)


def tekst(cel) -> str:
    """Celwaarde als gestripte tekst; een lege cel wordt een lege string.

    Neemt een cel (niet een waarde), zodat de aanroepen kort blijven. Getallen worden hun tekstvorm;
    een geheel getal dat als float uit Excel komt verliest zijn '.0', anders zou een nummer in de
    bron als '15.0' terechtkomen.
    """
    waarde = cel.value if hasattr(cel, "value") else cel
    if waarde is None:
        return ""
    if isinstance(waarde, str):
        return waarde.strip()
    if isinstance(waarde, float) and waarde.is_integer():
        return str(int(waarde))
    return str(waarde).strip()


def getal(cel) -> int | None:
    """Celwaarde als geheel getal, of None als er geen getal in staat.

    Het werkboek zet sommige getallen als tekst neer ('4' in plaats van 4). Een isinstance-controle
    op int zou die stil overslaan, dus we gaan via de tekstvorm.
    """
    waarde = tekst(cel)
    try:
        return int(float(waarde))
    except ValueError:
        return None


# De werkbooknamen die in een sjabloonformule mogen voorkomen, en waar ze in de tekst op uitkomen.
# De pagina vult ze in met de actuele objectnaam en het actuele niveau, net zoals Excel dat doet.
NAAM_PLAATSHOUDER = {"Object": "{object}", "Niveau": "{niveau}"}


def evalueer_sjabloon(formule: str, waar: str) -> str:
    """Een tekstformule van de vorm ="a"&"b"&Object&"c"&Niveau naar tekst met plaatshouders.

    Twee control-eisen op VSE zijn geen vaste tekst maar een sjabloon: de richtlijntekst met de
    objectnaam en het weerstandsniveau erin. Excel bewaart daarvan geen berekende waarde zolang het
    werkboek niet in Excel is geopend en opgeslagen, en die waarde zou hoe dan ook aan een object
    vastzitten. Daarom lezen we de formule zelf en houden we de plaatshouders open.
    """
    rest = formule[1:] if formule.startswith("=") else formule
    delen: list[str] = []
    i = 0
    while i < len(rest):
        if rest[i] == '"':
            j = i + 1
            stuk = []
            while j < len(rest):
                if rest[j] == '"':
                    if j + 1 < len(rest) and rest[j + 1] == '"':   # "" is een letterlijk aanhalingsteken
                        stuk.append('"')
                        j += 2
                        continue
                    break
                stuk.append(rest[j])
                j += 1
            else:
                sys.exit(f"{waar}: aanhalingsteken niet gesloten in de formule")
            delen.append("".join(stuk))
            i = j + 1
        elif rest[i] in " &":
            i += 1
        else:
            j = i
            while j < len(rest) and rest[j] not in '&" ':
                j += 1
            naam = rest[i:j].strip()
            if naam not in NAAM_PLAATSHOUDER:
                sys.exit(f"{waar}: onbekende verwijzing '{naam}' in de formule; "
                         f"bekend zijn {sorted(NAAM_PLAATSHOUDER)}")
            delen.append(NAAM_PLAATSHOUDER[naam])
            i = j
    return "".join(delen).strip()


def celtekst(ws_formules, rij: int, kolom: int, waar: str) -> tuple[str, bool]:
    """De tekst van een cel, en of het een sjabloon met plaatshouders is."""
    ruw = ws_formules.cell(rij, kolom).value
    if isinstance(ruw, str) and ruw.startswith("="):
        return evalueer_sjabloon(ruw, waar), True
    return tekst(ws_formules.cell(rij, kolom)), False


def sha256_bestand(pad: pathlib.Path) -> str:
    return hashlib.sha256(pad.read_bytes()).hexdigest()


def vingerafdruk(bron: dict) -> str:
    """Sha256 over de inhoud van de bron, niet over de bytes van het bestand.

    Git zet regeleindes op Windows om; hashen over bytes zou de controle op de ene machine laten
    slagen en op de andere niet, terwijl er inhoudelijk niets verschilt. Vast serialiseren met
    gesorteerde sleutels maakt de afdruk onafhankelijk van platform en volgorde.
    """
    kern = {sleutel: bron[sleutel] for sleutel in
            ("controls", "maatregelen", "bijlagen", "classificatie", "functiebox_niveau")}
    ruw = json.dumps(kern, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(ruw.encode("utf-8")).hexdigest()


def splits_aangeroepen(waarde: str, waar: str) -> list[str]:
    """Kolom 'Aangeroepen §2' naar een lijst paragraaf-ids.

    De cel bevat doorgaans een enkele verwijzing, maar het scheidingsteken is niet vastgelegd; we
    accepteren komma, puntkomma en schuine streep. Wat daarna geen paragraaf is, is een fout in de
    bron en geen reden om stil door te lopen.
    """
    if not waarde:
        return []
    delen = [d.strip() for d in re.split(r"[,;/]", waarde) if d.strip()]
    for deel in delen:
        if not PARAGRAAF.match(deel):
            sys.exit(f"{waar}: '{deel}' is geen paragraafverwijzing (verwacht §2.x of §2.x.y)")
    return delen


def splits_bijlagen(waarde: str) -> list[str]:
    """Kolom 'Bijlage (CSR)' naar bijlage-ids: '15' wordt 'CSR 15'; A, B en C blijven zoals ze zijn."""
    if not waarde:
        return []
    uit = []
    for deel in [d.strip() for d in re.split(r"[,;/]", waarde) if d.strip()]:
        uit.append(f"CSR {int(deel)}" if deel.isdigit() else deel)
    return uit


def lees_controls(wb, wb_formules) -> list[dict]:
    """De 127 controls: 89 op VSP Proceseisen (rij 4 t/m 92) en 38 op VSE Systeemeisen (rij 4 t/m 41)."""
    controls = []
    for blad, laatste in (("VSP Proceseisen", 92), ("VSE Systeemeisen", 41)):
        ws = wb[blad]
        ws_f = wb_formules[blad]
        kort_blad = blad.split()[0]
        for rij in range(4, laatste + 1):
            waar = f"{blad} rij {rij}"
            nr = getal(ws.cell(rij, 1))
            eis, sjabloon = celtekst(ws_f, rij, 3, waar)
            kort, _ = celtekst(ws_f, rij, 11, waar)
            if nr is None or not eis:
                sys.exit(f"{waar}: nummer of control-eis ontbreekt")
            controls.append({
                "id": f"{kort_blad}-{nr}",
                "blad": kort_blad,
                "nr": nr,
                "bio_bron": tekst(ws.cell(rij, 2)),
                "eis": eis,
                "sjabloon": sjabloon,
                "aangeroepen": splits_aangeroepen(tekst(ws.cell(rij, 4)), waar),
                "bijlagen": splits_bijlagen(tekst(ws.cell(rij, 5))),
                "kort": kort,
            })
    return controls


def lees_maatregelen(wb) -> list[dict]:
    """De 268 maatregelen (rij 4 t/m 271). Kolom F t/m I markeren met een X op welk niveau ze gelden."""
    ws = wb["Maatregelen"]
    maatregelen = []
    for rij in range(4, 272):
        code = tekst(ws.cell(rij, 4))
        if not code:
            sys.exit(f"Maatregelen rij {rij}: code ontbreekt")
        niveaus = []
        for nummer, kolom in enumerate((6, 7, 8, 9), start=1):
            merk = tekst(ws.cell(rij, kolom))
            if merk == "X":
                niveaus.append(nummer)
            elif merk:
                sys.exit(f"Maatregelen rij {rij} kolom {kolom}: '{merk}' verwacht was X of leeg")
        maatregelen.append({
            "code": code,
            "paragraaf": tekst(ws.cell(rij, 1)),
            "thema": tekst(ws.cell(rij, 2)),
            "groep": tekst(ws.cell(rij, 3)),
            "tekst": tekst(ws.cell(rij, 5)),
            "niveaus": niveaus,
        })
    return maatregelen


def lees_bijlagen(wb) -> list[dict]:
    """De 27 bijlagen (rij 4 t/m 30): CSR 1 t/m 24 plus A, B en C."""
    ws = wb["Bijlagen"]
    bijlagen = []
    for rij in range(4, 31):
        ident = tekst(ws.cell(rij, 1))
        if not ident:
            sys.exit(f"Bijlagen rij {rij}: id ontbreekt")
        bijlagen.append({
            "id": ident,
            "titel": tekst(ws.cell(rij, 2)),
            "type": tekst(ws.cell(rij, 3)),
            # Bijlage A wordt door geen enkele control aangeroepen; het werkboek zet daar een streepje.
            "aangeroepen_door": getal(ws.cell(rij, 4)),
        })
    return bijlagen


def lees_paragrafen(wb, maatregelen: list[dict]) -> list[dict]:
    """De 15 maatregelparagrafen, in de volgorde waarin ze op het maatregelenblad staan."""
    volgorde: list[str] = []
    themas: dict[str, str] = {}
    for maatregel in maatregelen:
        if maatregel["paragraaf"] not in themas:
            volgorde.append(maatregel["paragraaf"])
            themas[maatregel["paragraaf"]] = maatregel["thema"]
    return [{
        "id": par,
        "ouder": ouder(par),
        "thema": themas[par],
        "maatregelen": sum(1 for m in maatregelen if m["paragraaf"] == par),
    } for par in volgorde]


def ouder(paragraaf: str) -> str:
    """§2.5.1 hoort bij §2.5; §2.2 is zijn eigen ouder.

    Dit is letterlijk de verborgen hulpkolom Q van het maatregelenblad: heeft het nummer twee of meer
    punten, dan alles tot en met het tweede segment. Een control die §2.5 aanroept, raakt daarmee ook
    §2.5.1, §2.5.2 en §2.5.3.
    """
    if paragraaf.count(".") >= 2:
        eerste = paragraaf.index(".")
        tweede = paragraaf.index(".", eerste + 1)
        return paragraaf[:tweede]
    return paragraaf


def keuzelijst(wb_formules, blad: str, cel: str) -> list[str]:
    """De keuzelijst die op een cel ligt, uit de gegevensvalidatie van het werkboek.

    De pagina mag geen eigen lijstje statussen bijhouden: dan zou een wijziging in het werkboek stil
    langs de tool lopen. Daarom komen ook de keuzes uit de bron.
    """
    ws = wb_formules[blad]
    for validatie in ws.data_validations.dataValidation:
        if cel in validatie.sqref and validatie.formula1:
            return [d.strip() for d in validatie.formula1.strip('"').split(",") if d.strip()]
    sys.exit(f"{blad}!{cel}: geen keuzelijst gevonden")


def lees_keuzes(wb_formules) -> dict[str, list[str]]:
    return {
        "van_toepassing": keuzelijst(wb_formules, "VSP Proceseisen", "F4"),
        "status": keuzelijst(wb_formules, "VSP Proceseisen", "G4"),
        "functiebox": keuzelijst(wb_formules, "Instellingen", "C6"),
        "ja_nee": keuzelijst(wb_formules, "Instellingen", "C12"),
    }


def lees_functiebox_niveau(wb) -> dict[str, int]:
    """De tabel functiebox naar weerstandsniveau op blad Instellingen (B25:C29)."""
    ws = wb["Instellingen"]
    tabel = {}
    for rij in range(25, 30):
        box = tekst(ws.cell(rij, 2))
        niveau = getal(ws.cell(rij, 3))
        if box and niveau is not None:
            tabel[box] = niveau
    if sorted(tabel) != ["A", "B", "C", "D", "E"]:
        sys.exit(f"Instellingen B25:C29: verwacht A t/m E, gevonden {sorted(tabel)}")
    return tabel


def lees_classificatie() -> dict:
    """Het objectclassificatieformulier: zes criteria met vijf drempels, en de ernsttabel."""
    wb = openpyxl.load_workbook(CLASSIFICATIE, data_only=True)
    ws = wb.active

    ernst = []
    for rij in range(21, 26):
        score = getal(ws.cell(rij, 2))
        niveau = getal(ws.cell(rij, 19))             # kolom S
        if score is None or niveau is None:
            sys.exit(f"Objectclassificatie rij {rij}: score of niveau ontbreekt")
        ernst.append({
            "score": score,
            "label": tekst(ws.cell(rij, 3)),
            "functiebox": tekst(ws.cell(rij, 18)),   # kolom R
            "niveau": niveau,
        })

    criteria = []
    for ident, kolom_uitleg, kolom_drempel in CRITERIA:
        u = openpyxl.utils.column_index_from_string(kolom_uitleg)
        d = openpyxl.utils.column_index_from_string(kolom_drempel)
        criteria.append({
            "id": ident,
            "titel": tekst(ws.cell(19, u)),
            "uitleg": tekst(ws.cell(20, u)),
            "drempels": {str(rij - 20): tekst(ws.cell(rij, d)) for rij in range(21, 26)},
        })

    hoofdtaken = [tekst(ws.cell(rij, 8)) for rij in range(44, 52)]
    hoofdtaken = [h for h in hoofdtaken if h and h != VANGNET] + EXTRA_HOOFDTAKEN + [VANGNET]

    return {
        "ernst": ernst,
        "criteria": criteria,
        "hoofdtaken": hoofdtaken,
        "rekenregel_standaard": "gemiddelde",
        "toelichting_rekenregel": TOELICHTING_REKENREGEL,
    }


def bouw_bron() -> dict:
    for pad in (WERKBOEK, CLASSIFICATIE):
        if not pad.is_file():
            sys.exit(f"werkboek ontbreekt: {pad}")
    wb = openpyxl.load_workbook(WERKBOEK, data_only=True)
    wb_formules = openpyxl.load_workbook(WERKBOEK, data_only=False)
    maatregelen = lees_maatregelen(wb)
    return {
        "versie": VERSIE,
        "bron": {
            "richtlijn": "Cybersecurity Implementatierichtlijn Objecten (CSIR) 3.0, "
                         "definitief concept 14-09-2021",
            "gestripte_variant": "CSIR 3.4, gestripte variant voor aanbestedingen; hoofdstuk 2 en de "
                                 "bijlagen zijn identiek aan 3.0",
            "uitgever": "Rijkswaterstaat en Het Waterschapshuis",
            "auteursrecht": AUTEURSRECHT,
            "werkboek_versie": tekst(wb["Instellingen"].cell(38, 3)),
            "werkboek_sha256": sha256_bestand(WERKBOEK),
            "classificatie_sha256": sha256_bestand(CLASSIFICATIE),
            "gegenereerd_door": "register/haal_bron.py; wijzig de werkboeken, niet dit bestand met de hand",
        },
        "keuzes": lees_keuzes(wb_formules),
        "functiebox_niveau": lees_functiebox_niveau(wb),
        "classificatie": lees_classificatie(),
        "paragrafen": lees_paragrafen(wb, maatregelen),
        "controls": lees_controls(wb, wb_formules),
        "maatregelen": maatregelen,
        "bijlagen": lees_bijlagen(wb),
    }


def verschillen(nieuw: dict, oud: dict, pad: str = "") -> list[str]:
    """De eerste verschillen tussen twee bronnen, als leesbare paden."""
    uit: list[str] = []
    if type(nieuw) is not type(oud):
        return [f"{pad or 'wortel'}: {type(oud).__name__} werd {type(nieuw).__name__}"]
    if isinstance(nieuw, dict):
        for sleutel in sorted(set(nieuw) | set(oud)):
            if sleutel not in oud:
                uit.append(f"{pad}.{sleutel}: nieuw")
            elif sleutel not in nieuw:
                uit.append(f"{pad}.{sleutel}: verdwenen")
            else:
                uit += verschillen(nieuw[sleutel], oud[sleutel], f"{pad}.{sleutel}")
            if len(uit) >= 3:
                return uit[:3]
    elif isinstance(nieuw, list):
        if len(nieuw) != len(oud):
            return [f"{pad}: {len(oud)} werd {len(nieuw)} regels"]
        for i, (a, b) in enumerate(zip(nieuw, oud)):
            uit += verschillen(a, b, f"{pad}[{i}]")
            if len(uit) >= 3:
                return uit[:3]
    elif nieuw != oud:
        uit.append(f"{pad}: {oud!r} werd {nieuw!r}")
    return uit[:3]


def schrijf(bron: dict, doel: pathlib.Path) -> None:
    ruw = json.dumps(bron, ensure_ascii=False, indent=1) + "\n"
    doel.write_bytes(ruw.encode("utf-8"))


def main(argv: list[str]) -> int:
    bron = bouw_bron()
    if "--check" in argv:
        if not DOEL.is_file():
            print(f"{DOEL.name} ontbreekt; draai dit script zonder --check.")
            return 1
        oud = json.loads(DOEL.read_text(encoding="utf-8"))
        if oud == bron:
            print(f"{DOEL.name} loopt gelijk met de werkboeken.")
            return 0
        print(f"{DOEL.name} wijkt af van de werkboeken:")
        for regel in verschillen(bron, oud):
            print(f"  {regel}")
        print("Draai 'python register/haal_bron.py' om hem bij te werken.")
        return 1
    schrijf(bron, DOEL)
    print(f"{DOEL}: {len(bron['controls'])} controls, {len(bron['maatregelen'])} maatregelen, "
          f"{len(bron['bijlagen'])} bijlagen, {len(bron['paragrafen'])} paragrafen, "
          f"vingerafdruk {vingerafdruk(bron)[:12]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
